from pprint import pprint
import openpyxl as xl

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
cell = sheet.cell(row=2, column=2)
rows = sheet.max_row
columns = sheet.max_column
cell_value = cell.value
print(cell_value)
for row in range(2, rows +1):
    for column in range(1, columns + 1):
        cell_value = sheet.cell(row=row, column=column).value
        pprint(cell_value, width=50)

# sheet.insert_rows()
# sheet.insert_cols()
# sheet.append()
# sheet.delete_rows()
# sheet.delete_cols()
